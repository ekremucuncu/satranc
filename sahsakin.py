import beyazpiyon
import beyazkale
import beyazfil
import beyazvezir
import beyazsah
import beyazat
import siyahpiyon
import siyahkale
import siyahfil
import siyahvezir
import siyahsah
import siyahat
from openpyxl import *


def beyaz_sira(sheet,kitap):
    b = 0
    for row in sheet.iter_rows(min_row=2, max_row=33, min_col=1, max_col=1):
        for cell in row:
            if cell.value != None:
                if "Siyah" in cell.value:
                    b = b + 1
    liste=beyaz_sira_degerli(sheet,kitap,b)
    return liste


def beyaz_sira_degerli(sheet,kitap,b):
    liste=[]
    for row in sheet.iter_rows(min_row=2, max_row=33, min_col=1, max_col=1):
        for cell in row:
            if cell.value != None:
                if "Beyaz Sah" in cell.value:
                    kitap = load_workbook("satranc.xlsx")
                    sheet = kitap.active
                    a=tasa_baglanma_siyah_61(sheet.cell(cell.row, 2), sheet.cell(cell.row, 3), cell.value,sheet)
                    for i in range(len(a)):
                        aaa=tas_hareket_ettirme_siyah_61(a, sheet.cell(cell.row, 2).value,sheet.cell(cell.row, 3).value, sheet, i)
                        #burada beyaz tasi hareket ettirdik
                        tutucu,x,y=beyaz_sira_deneme(sheet,sheet.cell(cell.row,2).value,sheet.cell(cell.row,3).value,kitap,b,aaa)
                        if tutucu==0:
                            liste.append((cell.value,x,y))
                        kitap = load_workbook("satranc.xlsx")
                        sheet = kitap.active
    return liste


def beyaz_sira_deneme(sheet,x_jesus,y_jesus,kitap,b,aaa):
    isa=[]
    a=0
    c=0
    for row in sheet.iter_rows(min_row=2, max_row=33, min_col=1, max_col=1):
        for cell in row:
            if cell.value != None:
                if "Siyah" in cell.value:
                    deger,x,y=tasa_baglanma_siyah(sheet.cell(cell.row, 2), sheet.cell(cell.row, 3), cell.value,sheet,x_jesus,y_jesus)
                    isa.append(deger)
                    c=c+1
    if b!=len(isa):
        isa.append(0)
    for i in range(len(isa)):
        if isa[i]==0:
            a=a+1
    if b==c:
        if a==len(isa):
            return 1,x,y
        if a!=len(isa):
            return 0,x,y
    else:
        x=aaa[0]
        y=aaa[1]
        if a==len(isa):
            return 1,x,y
        if a!=len(isa):
            return 0,x,y


def tasa_baglanma_siyah_61(x_secilen,y_secilen,secilen_tas,sheet):
    if f"Beyaz Sah" in secilen_tas:
        a,b=beyazsah.beyazsah(x_secilen,y_secilen,secilen_tas,sheet)
    return a


def tas_hareket_ettirme_siyah_61(a,x_bulunan,y_bulunan,sheet,j):
    gidilecek_yer = a[j]
    aaa=tas_yok_etme(gidilecek_yer[0], gidilecek_yer[1], sheet)
    for row in sheet.iter_rows(min_row=2, max_row=33, min_col=2, max_col=2):
        for cell in row:
            if cell.value == x_bulunan:
                y = sheet.cell(cell.row, 3)
                if y.value == y_bulunan:
                    isim = sheet.cell(cell.row, 1).value
                    sheet.cell(row=cell.row, column=2, value=gidilecek_yer[0])
                    sheet.cell(row=cell.row, column=3, value=gidilecek_yer[1])
    return aaa


def tas_yok_etme(x_olen, y_olen, sheet):
    for row in sheet.iter_rows(min_row=2, max_row=33, min_col=2, max_col=2):
        for cell in row:
            if cell.value == x_olen:
                y = sheet.cell(cell.row, 3)
                isim = sheet.cell(cell.row, 1).value
                if y.value == y_olen:
                    sheet.delete_rows(cell.row)
                    return x_olen,y_olen


def tasa_baglanma_siyah(x_secilen,y_secilen,secilen_tas,sheet,x_jesus,y_jesus):
    for x in range(1,11):
        if f"Siyah Piyon{x}" in secilen_tas:
            a,b=siyahpiyon.siyahpiyon(x_secilen,y_secilen,secilen_tas,sheet)
            break
        if f"Siyah Kale{x}" in secilen_tas:
            a,b=siyahkale.siyahkale(x_secilen,y_secilen,secilen_tas,sheet)
            break
        if f"Siyah Fil{x}" in secilen_tas:
            a,b=siyahfil.siyahfil(x_secilen,y_secilen,secilen_tas,sheet)
            break
        if f"Siyah Vezir{x}" in secilen_tas:
            a,b=siyahvezir.siyahvezir(x_secilen,y_secilen,secilen_tas,sheet)
            break
        if f"Siyah Sah" in secilen_tas:
            a,b=siyahsah.siyahsah(x_secilen,y_secilen,secilen_tas,sheet)
            break
        if f"Siyah At{x}" in secilen_tas:
            a,b=siyahat.siyahat(x_secilen,y_secilen,secilen_tas,sheet)
            break
    deger =siyah(a,sheet,x_jesus,y_jesus)
    return deger


def siyah(a,sheet,x_jesus,y_jesus):
    for row in sheet.iter_rows(min_row=2, max_row=33, min_col=1, max_col=1):
        for cell in row:
            if cell.value != None:
                if "Beyaz Sah" in cell.value:
                    x_sah=sheet.cell(cell.row,2).value
                    y_sah=sheet.cell(cell.row,3).value
                    konum=(x_sah,y_sah)
                    if konum in a and x_jesus==0 and y_jesus==0:
                        return 61,x_sah,y_sah
                    if konum in a and x_jesus!=0 and y_jesus!=0:
                        return 61,x_jesus,y_jesus
                    if konum not in a and x_jesus==0 and y_jesus==0:
                        return 0,x_sah,y_sah
                    if konum not in a and x_jesus!=0 and y_jesus!=0:
                        return 0,x_jesus,y_jesus

#--------------------------------------------------------------

def siyah_sira(sheet,kitap):
    b=0
    for row in sheet.iter_rows(min_row=2, max_row=33, min_col=1, max_col=1):
        for cell in row:
            if cell.value != None:
                if "Beyaz" in cell.value:
                    b=b+1
    liste=siyah_sira_degerli(sheet,kitap,b)
    return liste


def siyah_sira_deneme(sheet,x_jesus,y_jesus,kitap,b,aaa):
    isa=[]
    a=0
    c=0
    for row in sheet.iter_rows(min_row=2, max_row=33, min_col=1, max_col=1):
        for cell in row:
            if cell.value != None:
                if "Beyaz" in cell.value:
                    deger,x,y=tasa_baglanma_beyaz(sheet.cell(cell.row, 2), sheet.cell(cell.row, 3), cell.value,sheet,x_jesus,y_jesus)
                    isa.append(deger)
                    c=c+1
    if b!=len(isa):
        isa.append(0)
    for i in range(len(isa)):
        if isa[i]==0:
            a=a+1
    if b==c:
        if a==len(isa):
            return 1,x,y
        if a!=len(isa):
            return 0,x,y
    else:
        x=aaa[0]
        y=aaa[1]
        if a==len(isa):
            return 1,x,y
        if a!=len(isa):
            return 0,x,y


def siyah_sira_degerli(sheet,kitap,b):
    liste=[]
    for row in sheet.iter_rows(min_row=2, max_row=33, min_col=1, max_col=1):
        for cell in row:
            if cell.value != None:
                if "Siyah Sah" in cell.value:
                    kitap = load_workbook("satranc.xlsx")
                    sheet = kitap.active
                    a=tasa_baglanma_beyaz_61(sheet.cell(cell.row, 2), sheet.cell(cell.row, 3), cell.value,sheet)
                    for i in range(len(a)):
                        aaa=tas_hareket_ettirme_beyaz_61(a, sheet.cell(cell.row, 2).value,sheet.cell(cell.row, 3).value, sheet, i)
                        #burada siyah tasi hareket ettirdik
                        tutucu,x,y=siyah_sira_deneme(sheet,sheet.cell(cell.row,2).value,sheet.cell(cell.row,3).value,kitap,b,aaa)
                        if tutucu==0:
                            liste.append((cell.value,x,y))
                        kitap = load_workbook("satranc.xlsx")
                        sheet = kitap.active
    return liste


def tasa_baglanma_beyaz(x_secilen,y_secilen,secilen_tas,sheet,x_jesus,y_jesus):
    for x in range(1,11):
        if f"Beyaz Piyon{x}" in secilen_tas:
            a,b=beyazpiyon.beyazpiyon(x_secilen,y_secilen,secilen_tas,sheet)
            break
        if f"Beyaz Kale{x}" in secilen_tas:
            a,b=beyazkale.beyazkale(x_secilen,y_secilen,secilen_tas,sheet)
            break
        if f"Beyaz Fil{x}" in secilen_tas:
            a,b=beyazfil.beyazfil(x_secilen,y_secilen,secilen_tas,sheet)
            break
        if f"Beyaz Vezir{x}" in secilen_tas:
            a,b=beyazvezir.beyazvezir(x_secilen,y_secilen,secilen_tas,sheet)
            break
        if f"Beyaz Sah" in secilen_tas:
            a,b=beyazsah.beyazsah(x_secilen,y_secilen,secilen_tas,sheet)
    deger =beyaz(a,sheet,x_jesus,y_jesus)
    return deger


def tasa_baglanma_beyaz_61(x_secilen,y_secilen,secilen_tas,sheet):
    if f"Siyah Sah" in secilen_tas:
        a,b=siyahsah.siyahsah(x_secilen,y_secilen,secilen_tas,sheet)
    return a


def tas_hareket_ettirme_beyaz_61(a,x_bulunan,y_bulunan,sheet,j):
    gidilecek_yer = a[j]
    aaa=tas_yok_etme(gidilecek_yer[0], gidilecek_yer[1], sheet)
    for row in sheet.iter_rows(min_row=2, max_row=33, min_col=2, max_col=2):
        for cell in row:
            if cell.value == x_bulunan:
                y = sheet.cell(cell.row, 3)
                if y.value == y_bulunan:
                    isim = sheet.cell(cell.row, 1).value
                    sheet.cell(row=cell.row, column=2, value=gidilecek_yer[0])
                    sheet.cell(row=cell.row, column=3, value=gidilecek_yer[1])
    return aaa


def tas_yok_etme(x_olen, y_olen, sheet):
    for row in sheet.iter_rows(min_row=2, max_row=33, min_col=2, max_col=2):
        for cell in row:
            if cell.value == x_olen:
                y = sheet.cell(cell.row, 3)
                isim = sheet.cell(cell.row, 1).value
                if y.value == y_olen:
                    sheet.delete_rows(cell.row)
                    return x_olen,y_olen


def beyaz(a,sheet,x_jesus,y_jesus):
    for row in sheet.iter_rows(min_row=2, max_row=33, min_col=1, max_col=1):
        for cell in row:
            if cell.value != None:
                if "Siyah Sah" in cell.value:
                    x_sah=sheet.cell(cell.row,2).value
                    y_sah=sheet.cell(cell.row,3).value
                    konum=(x_sah,y_sah)
                    if konum in a and x_jesus==0 and y_jesus==0:
                        return 61,x_sah,y_sah
                    if konum in a and x_jesus!=0 and y_jesus!=0:
                        return 61,x_jesus,y_jesus
                    if konum not in a and x_jesus==0 and y_jesus==0:
                        return 0,x_sah,y_sah
                    if konum not in a and x_jesus!=0 and y_jesus!=0:
                        return 0,x_jesus,y_jesus