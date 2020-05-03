import sahikoruma
import saholuyor
import sahsakin
import beyaz
import siyah
from openpyxl import *


def beyaz_sira():
    kitap = load_workbook("satranc.xlsx")
    sheet = kitap.active
    ekrem=sahikoruma.beyaz_sira(sheet)
    liste1,sahmatmi=saholuyor.beyaz_sira(sheet,kitap)
    kitap = load_workbook("satranc.xlsx")
    sheet = kitap.active
    liste2=sahsakin.beyaz_sira(sheet,kitap)
    liste3=remove(liste1+liste2)
    kitap = load_workbook("satranc.xlsx")
    sheet = kitap.active
    if liste1!=[]:
        print("Sah")
        if sahmatmi==[]:
            print("Sah Mat")
            print("Yenen Siyah")
            exit()
    print("Sira Beyazda")
    beyaz.tas_secme(sheet,ekrem,liste3)
    kitap.save("satranc.xlsx")


def siyah_sira():
    kitap = load_workbook("satranc.xlsx")
    sheet = kitap.active
    ekrem=sahikoruma.siyah_sira(sheet)
    liste1,sahmatmi=saholuyor.siyah_sira(sheet,kitap)
    kitap=load_workbook("satranc.xlsx")
    sheet=kitap.active
    liste2=sahsakin.siyah_sira(sheet,kitap)
    liste3=remove(liste1+liste2)
    kitap = load_workbook("satranc.xlsx")
    sheet = kitap.active
    if liste1!=[]:
        print("Sah")
        if sahmatmi==[]:
            print("Sah Mat")
            print("Yenen Beyaz")
            exit()
    print("Sira Siyahta")
    siyah.tas_secme(sheet,ekrem,liste3)
    kitap.save("satranc.xlsx")


def remove(duplicate):
    final_list = []
    for num in duplicate:
        if num not in final_list:
            final_list.append(num)
    return final_list


print("Hosgeldiniz")
sira=2
while True:
    if sira%2==0:
        sira=sira+1
        beyaz_sira()
    if sira%2==1:
        sira=sira+1
        siyah_sira()


