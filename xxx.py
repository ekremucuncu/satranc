from openpyxl import *


kitap1 = load_workbook("satranc.xlsx")
sheet1 = kitap1.active
kitap2 = load_workbook("satranc_degismeyen.xlsx")
sheet2=kitap2.active

for row in sheet2.iter_rows(min_col=1,max_col=3,min_row=1,max_row=33):
    for cell in row:
        a=cell.value
        sheet1.cell(cell.row,cell.column,cell.value)
kitap1.save("satranc.xlsx")